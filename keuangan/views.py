from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.utils import timezone
import json
from django.http import HttpResponse
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image
)
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test

from .models import *

def is_admin(user):
    return user.groups.filter(name='Admin').exists()

def is_direktur(user):
    return user.groups.filter(name='Direktur').exists()

def login_view(request):

    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':

        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(
            request,
            username=username,
            password=password
        )

        if user is not None:

            login(request, user)

            return redirect('dashboard')

        messages.error(
            request,
            'Username atau Password salah'
        )

    return render(
        request,
        'login.html'
    )

# Create your views here.
@login_required(login_url='login')
def dashboard(request):

    total_masuk = (
        KasMasuk.objects.aggregate(
            total=Sum('jumlah')
        )['total'] or 0
    )

    total_keluar = (
        KasKeluar.objects.aggregate(
            total=Sum('jumlah')
        )['total'] or 0
    )

    saldo = total_masuk - total_keluar

    total_transaksi = (
        KasMasuk.objects.count() +
        KasKeluar.objects.count()
    )

    kas_masuk_terbaru = (
        KasMasuk.objects
        .order_by('-tanggal', '-jam')[:5]
    )

    kas_keluar_terbaru = (
        KasKeluar.objects
        .order_by('-tanggal', '-jam')[:5]
    )

    # Statistik tambahan

    hari_ini = timezone.now().date()

    kas_masuk_hari_ini = (
        KasMasuk.objects.filter(
            tanggal=hari_ini
        ).aggregate(
            total=Sum('jumlah')
        )['total'] or 0
    )

    kas_keluar_hari_ini = (
        KasKeluar.objects.filter(
            tanggal=hari_ini
        ).aggregate(
            total=Sum('jumlah')
        )['total'] or 0
    )

    jumlah_karyawan = Karyawan.objects.count()

    # Grafik bulanan

    data_masuk = (
        KasMasuk.objects
        .annotate(bulan=TruncMonth('tanggal'))
        .values('bulan')
        .annotate(total=Sum('jumlah'))
        .order_by('bulan')
    )

    data_keluar = (
        KasKeluar.objects
        .annotate(bulan=TruncMonth('tanggal'))
        .values('bulan')
        .annotate(total=Sum('jumlah'))
        .order_by('bulan')
    )

    labels = []
    chart_masuk = []
    chart_keluar = []

    bulan_dict = {}

    for item in data_masuk:
        bulan = item['bulan'].strftime('%b %Y')
        bulan_dict[bulan] = {
            'masuk': float(item['total']),
            'keluar': 0
        }

    for item in data_keluar:
        bulan = item['bulan'].strftime('%b %Y')

        if bulan not in bulan_dict:
            bulan_dict[bulan] = {
                'masuk': 0,
                'keluar': float(item['total'])
            }
        else:
            bulan_dict[bulan]['keluar'] = float(item['total'])

    for bulan, nilai in bulan_dict.items():
        labels.append(bulan)
        chart_masuk.append(nilai['masuk'])
        chart_keluar.append(nilai['keluar'])

    context = {
        'total_masuk': total_masuk or 0,
        'total_keluar': total_keluar or 0,
        'saldo': saldo or 0,
        'total_transaksi': total_transaksi or 0,

        'kas_masuk_terbaru': kas_masuk_terbaru,
        'kas_keluar_terbaru': kas_keluar_terbaru,

        'jumlah_karyawan': jumlah_karyawan or 0,
        'kas_masuk_hari_ini': kas_masuk_hari_ini or 0,
        'kas_keluar_hari_ini': kas_keluar_hari_ini or 0,

        'chart_labels': labels,
        'chart_masuk': chart_masuk,
        'chart_keluar': chart_keluar,
    }

    return render(
        request,
        'dashboard.html',
        context
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def kas_masuk(request):
    data = KasMasuk.objects.all().order_by('-tanggal')
    return render(request, 'kas_masuk/list.html', {
        'data': data
    })

def tambah_kas_masuk(request):

    if request.method == 'POST':

        KasMasuk.objects.create(
            tanggal=request.POST['tanggal'],
            jam=request.POST['jam'],
            keterangan=request.POST['keterangan'],
            jumlah=request.POST['jumlah'],
            users=request.user.username
        )

        return redirect('kas_masuk')

    return render(
        request,
        'kas_masuk/create.html'
    )

def edit_kas_masuk(request, id):

    kas = get_object_or_404(
        KasMasuk,
        id=id
    )

    if request.method == 'POST':

        kas.tanggal = request.POST['tanggal']
        kas.jam = request.POST['jam']
        kas.keterangan = request.POST['keterangan']
        kas.jumlah = request.POST['jumlah']
        kas.users = request.user.username

        kas.save()

        return redirect('kas_masuk')

    return render(
        request,
        'kas_masuk/edit.html',
        {
            'kas': kas
        }
    )

def hapus_kas_masuk(request, id):

    data = get_object_or_404(
        KasMasuk,
        id=id
    )

    data.delete()

    return redirect('kas_masuk')

@login_required(login_url='login')
@user_passes_test(is_admin)
def kas_keluar(request):

    data = KasKeluar.objects.all()

    tanggal_awal = request.GET.get('tanggal_awal')
    tanggal_akhir = request.GET.get('tanggal_akhir')

    if tanggal_awal and tanggal_akhir:
        data = data.filter(
            tanggal__range=[
                tanggal_awal,
                tanggal_akhir
            ]
        )

    data = data.order_by('-id')

    return render(
        request,
        'kas_keluar/list.html',
        {
            'data': data,
            'tanggal_awal': tanggal_awal,
            'tanggal_akhir': tanggal_akhir
        }
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def tambah_kas_keluar(request):

    if request.method == 'POST':

        KasKeluar.objects.create(
            tanggal=request.POST['tanggal'],
            jam=request.POST['jam'],
            keterangan=request.POST['keterangan'],
            jumlah=request.POST['jumlah'],
            users=request.user.username
        )

        return redirect('kas_keluar')

    return render(
        request,
        'kas_keluar/create.html'
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def edit_kas_keluar(request, id):

    kas = get_object_or_404(
        KasKeluar,
        id=id
    )

    if request.method == 'POST':

        kas.tanggal = request.POST['tanggal']
        kas.jam = request.POST['jam']
        kas.keterangan = request.POST['keterangan']
        kas.jumlah = request.POST['jumlah']
        kas.save()

        return redirect('kas_keluar')

    return render(
        request,
        'kas_keluar/edit.html',
        {
            'kas': kas
        }
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def hapus_kas_keluar(request, id):

    data = get_object_or_404(
        KasKeluar,
        id=id
    )

    data.delete()

    return redirect('kas_keluar')


def dashboard(request):

    total_masuk = KasMasuk.objects.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    total_keluar = KasKeluar.objects.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    saldo = total_masuk - total_keluar

    total_transaksi = (
        KasMasuk.objects.count() +
        KasKeluar.objects.count()
    )

    kas_masuk_terbaru = KasMasuk.objects.order_by('-id')[:5]
    kas_keluar_terbaru = KasKeluar.objects.order_by('-id')[:5]

    context = {
        'total_masuk': total_masuk,
        'total_keluar': total_keluar,
        'saldo': saldo,
        'total_transaksi': total_transaksi,
        'kas_masuk_terbaru': kas_masuk_terbaru,
        'kas_keluar_terbaru': kas_keluar_terbaru,
    }

    return render(
        request,
        'dashboard.html',
        context
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def list_karyawan(request):
    karyawan = Karyawan.objects.all().order_by('-id')

    return render(
        request,
        'karyawan/list.html',
        {
            'karyawan': karyawan
        }
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def tambah_karyawan(request):

    if request.method == 'POST':
        Karyawan.objects.create(
            nama=request.POST['nama'],
            jabatan=request.POST['jabatan'],
            tgl_bergabung=request.POST['tgl_bergabung']
        )

        return redirect('list_karyawan')

    return render(request, 'karyawan/create.html')

@login_required(login_url='login')
@user_passes_test(is_admin)
def edit_karyawan(request, id):

    data = get_object_or_404(Karyawan, id=id)

    if request.method == 'POST':
        data.nama = request.POST['nama']
        data.jabatan = request.POST['jabatan']
        data.tgl_bergabung = request.POST['tgl_bergabung']
        data.save()

        return redirect('list_karyawan')

    return render(
        request,
        'karyawan/edit.html',
        {
            'data': data
        }
    )

@login_required(login_url='login')
@user_passes_test(is_admin)
def hapus_karyawan(request, id):

    data = get_object_or_404(Karyawan, id=id)
    data.delete()

    return redirect('list_karyawan')


@login_required(login_url='login')
def buku_besar(request):

    transaksi = []

    for item in KasMasuk.objects.all():

        transaksi.append({
            'tanggal': item.tanggal,
            'jam': item.jam,
            'keterangan': item.keterangan,
            'debit': item.jumlah,
            'kredit': 0,
            'users': item.users,
        })

    for item in KasKeluar.objects.all():

        transaksi.append({
            'tanggal': item.tanggal,
            'jam': item.jam,
            'keterangan': item.keterangan,
            'debit': 0,
            'kredit': item.jumlah,
            'users': item.users,
        })

    transaksi.sort(
        key=lambda x: (
            x['tanggal'],
            x['jam']
        )
    )

    saldo = 0

    for item in transaksi:

        saldo += item['debit']
        saldo -= item['kredit']

        item['saldo'] = saldo

    total_debit = KasMasuk.objects.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    total_kredit = KasKeluar.objects.aggregate(
        total=Sum('jumlah')
    )['total'] or 0

    saldo_akhir = total_debit - total_kredit

    total_transaksi = len(transaksi)

    return render(
        request,
        'buku_besar/list.html',
        {
            'transaksi': transaksi,
            'total_debit': total_debit,
            'total_kredit': total_kredit,
            'saldo_akhir': saldo_akhir,
            'total_transaksi': total_transaksi,
        }
    )


@login_required(login_url='login')
def neraca_saldo(request):

    total_debit = (
        KasMasuk.objects.aggregate(
            total=Sum('jumlah')
        )['total'] or 0
    )

    total_kredit = (
        KasKeluar.objects.aggregate(
            total=Sum('jumlah')
        )['total'] or 0
    )

    saldo = total_debit - total_kredit

    akun = [
        {
            'nama_akun': 'Kas',
            'debit': total_debit,
            'kredit': total_kredit,
            'saldo': saldo
        }
    ]

    context = {
        'akun': akun,
        'total_debit': total_debit,
        'total_kredit': total_kredit,
        'saldo': saldo,
    }

    return render(
        request,
        'neraca_saldo/list.html',
        context
    )

def tambah_kop(elements, styles):

    logo = Image(
        "static/assets/images/logo.jpeg",
        width=1.8*cm,
        height=1.8*cm
    )

    teks_kop = [
        Paragraph(
            "<para align='center'><font size='16'><b>WISATA ALAM BANGA</b></font></para>",
            styles['Normal']
        ),
        Spacer(1, 5),
        Paragraph(
            "<para align='center'><font size='10'>Desa Gattareng Toa, Kabupaten Soppeng</font></para>",
            styles['Normal']
        ),
        Spacer(1, 3),
        Paragraph(
            "<para align='center'><font size='10'>Sistem Informasi Akuntansi</font></para>",
            styles['Normal']
        ),
    ]

    kop = Table([
        [
            logo,
            teks_kop,
            ""
        ]
    ], colWidths=[2.5*cm, 11*cm, 2.5*cm])

    kop.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('LINEBELOW', (0, 0), (-1, -1), 1, colors.black),
    ]))

    elements.append(kop)
    elements.append(Spacer(1, 20))


def export_pdf_kas_keluar(request):

    tanggal_awal = request.GET.get('tanggal_awal')
    tanggal_akhir = request.GET.get('tanggal_akhir')

    data_queryset = KasKeluar.objects.all()

    if tanggal_awal and tanggal_akhir:
        data_queryset = data_queryset.filter(
            tanggal__range=[
                tanggal_awal,
                tanggal_akhir
            ]
        )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        'attachment; filename="laporan_kas_keluar.pdf"'
    )

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []

    tambah_kop(elements, styles)

    elements.append(
        Paragraph(
            "<b>LAPORAN KAS KELUAR</b>",
            styles['Heading1']
        )
    )

    if tanggal_awal and tanggal_akhir:
        elements.append(
            Paragraph(
                f"Periode : {tanggal_awal} s/d {tanggal_akhir}",
                styles['Normal']
            )
        )

    elements.append(Spacer(1, 15))

    table_data = [
        ['No', 'Tanggal', 'Jam', 'Keterangan', 'User', 'Jumlah']
    ]

    total = 0

    for no, item in enumerate(data_queryset, start=1):
        total += item.jumlah

        table_data.append([
            no,
            str(item.tanggal),
            str(item.jam),
            item.keterangan,
            item.users,
            f"Rp {item.jumlah:,.0f}"
        ])

    table_data.append([
        '',
        '',
        '',
        '',
        'TOTAL',
        f"Rp {total:,.0f}"
    ])

    table = Table(
        table_data,
        colWidths=[1*cm, 2.5*cm, 2*cm, 4.5*cm, 2*cm, 3.2*cm]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (2, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            f"Total Transaksi : {data_queryset.count()}",
            styles['Normal']
        )
    )

    doc.build(elements)

    return response

def export_excel_kas_keluar(request):

    tanggal_awal = request.GET.get('tanggal_awal')
    tanggal_akhir = request.GET.get('tanggal_akhir')

    data_queryset = KasKeluar.objects.all()

    if tanggal_awal and tanggal_akhir:
        data_queryset = data_queryset.filter(
            tanggal__range=[
                tanggal_awal,
                tanggal_akhir
            ]
        )

    wb = Workbook()
    ws = wb.active

    ws.title = "Kas Keluar"

    ws.append([
        'Tanggal',
        'Jam',
        'Keterangan',
        'Jumlah',
        'User'
    ])

    total = 0

    for item in data_queryset:

        total += item.jumlah

        ws.append([
            str(item.tanggal),
            str(item.jam),
            item.keterangan,
            float(item.jumlah),
            item.users
        ])

    ws.append([])
    ws.append([
        '',
        '',
        'TOTAL',
        float(total),
        ''
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=kas_keluar.xlsx'
    )

    wb.save(response)

    return response

def export_excel_kas_masuk(request):

    tanggal_awal = request.GET.get('tanggal_awal')
    tanggal_akhir = request.GET.get('tanggal_akhir')

    data_queryset = KasMasuk.objects.all()

    if tanggal_awal and tanggal_akhir:
        data_queryset = data_queryset.filter(
            tanggal__range=[
                tanggal_awal,
                tanggal_akhir
            ]
        )

    wb = Workbook()
    ws = wb.active

    ws.title = "Kas Masuk"

    ws.append([
        'Tanggal',
        'Jam',
        'Keterangan',
        'Jumlah',
        'User'
    ])

    total = 0

    for item in data_queryset:

        total += item.jumlah

        ws.append([
            str(item.tanggal),
            str(item.jam),
            item.keterangan,
            float(item.jumlah),
            item.users
        ])

    ws.append([])
    ws.append([
        '',
        '',
        'TOTAL',
        float(total),
        ''
    ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename=kas_masuk.xlsx'
    )

    wb.save(response)

    return response

def export_pdf_kas_masuk(request):

    tanggal_awal = request.GET.get('tanggal_awal')
    tanggal_akhir = request.GET.get('tanggal_akhir')

    data_queryset = KasMasuk.objects.all()

    if tanggal_awal and tanggal_akhir:
        data_queryset = data_queryset.filter(
            tanggal__range=[
                tanggal_awal,
                tanggal_akhir
            ]
        )

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = (
        'attachment; filename="laporan_kas_masuk.pdf"'
    )

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []

    # KOP LAPORAN
    tambah_kop(elements, styles)

    # JUDUL LAPORAN
    elements.append(
        Paragraph(
            "<b>LAPORAN KAS MASUK</b>",
            styles['Heading1']
        )
    )

    if tanggal_awal and tanggal_akhir:
        elements.append(
            Paragraph(
                f"Periode : {tanggal_awal} s/d {tanggal_akhir}",
                styles['Normal']
            )
        )

    elements.append(Spacer(1, 15))

    table_data = [
        ['No', 'Tanggal', 'Jam', 'Keterangan', 'User', 'Jumlah']
    ]

    total = 0

    for no, item in enumerate(data_queryset, start=1):

        total += item.jumlah

        table_data.append([
            no,
            str(item.tanggal),
            str(item.jam),
            item.keterangan,
            item.users,
            f"Rp {item.jumlah:,.0f}"
        ])

    table_data.append([
        '',
        '',
        '',
        '',
        'TOTAL',
        f"Rp {total:,.0f}"
    ])

    table = Table(
        table_data,
        colWidths=[1*cm, 2.5*cm, 2*cm, 4.5*cm, 2*cm, 3.2*cm]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (2, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            f"Total Transaksi : {data_queryset.count()}",
            styles['Normal']
        )
    )

    doc.build(elements)

    return response

def logout_view(request):

    logout(request)

    return redirect('login')

def export_pdf_neraca_saldo(request):

    total_debit = KasMasuk.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    total_kredit = KasKeluar.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    saldo = total_debit - total_kredit

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="neraca_saldo.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []

    # KOP LAPORAN
    tambah_kop(elements, styles)

    # JUDUL LAPORAN
    elements.append(
        Paragraph(
            "<b>NERACA SALDO</b>",
            styles['Heading1']
        )
    )

    elements.append(Spacer(1, 15))

    table_data = [
        ['No', 'Nama Akun', 'Debit', 'Kredit', 'Saldo'],
        [
            1,
            'Kas',
            f"Rp {total_debit:,.0f}",
            f"Rp {total_kredit:,.0f}",
            f"Rp {saldo:,.0f}"
        ]
    ]

    table = Table(
        table_data,
        colWidths=[1*cm, 5*cm, 3*cm, 3*cm, 3*cm]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            f"<b>Saldo Akhir : Rp {saldo:,.0f}</b>",
            styles['Normal']
        )
    )

    doc.build(elements)

    return response


def export_excel_neraca_saldo(request):
    total_debit = KasMasuk.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    total_kredit = KasKeluar.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    saldo = total_debit - total_kredit

    wb = Workbook()
    ws = wb.active
    ws.title = "Neraca Saldo"

    ws.append(['No', 'Nama Akun', 'Debit', 'Kredit', 'Saldo'])
    ws.append([1, 'Kas', float(total_debit), float(total_kredit), float(saldo)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=neraca_saldo.xlsx'

    wb.save(response)
    return response

@login_required(login_url='login')
def laporan_laba_rugi(request):
    total_pendapatan = KasMasuk.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    total_beban = KasKeluar.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    laba_rugi = total_pendapatan - total_beban

    return render(request, 'laba_rugi/list.html', {
        'total_pendapatan': total_pendapatan,
        'total_beban': total_beban,
        'laba_rugi': laba_rugi,
    })


def export_pdf_laba_rugi(request):

    total_pendapatan = KasMasuk.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    total_beban = KasKeluar.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    laba_rugi = total_pendapatan - total_beban

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="laporan_laba_rugi.pdf"'

    doc = SimpleDocTemplate(response)
    styles = getSampleStyleSheet()
    elements = []

    # KOP LAPORAN
    tambah_kop(elements, styles)

    # JUDUL LAPORAN
    elements.append(
        Paragraph(
            "<b>LAPORAN LABA RUGI</b>",
            styles['Heading1']
        )
    )

    elements.append(Spacer(1, 15))

    table_data = [
        ['Keterangan', 'Jumlah'],
        ['Pendapatan / Kas Masuk', f"Rp {total_pendapatan:,.0f}"],
        ['Beban / Kas Keluar', f"Rp {total_beban:,.0f}"],
        ['Laba / Rugi', f"Rp {laba_rugi:,.0f}"],
    ]

    table = Table(
        table_data,
        colWidths=[10*cm, 5*cm]
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 20))

    elements.append(
        Paragraph(
            f"Laba / Rugi Bersih : Rp {laba_rugi:,.0f}",
            styles['Normal']
        )
    )

    doc.build(elements)

    return response


def export_excel_laba_rugi(request):
    total_pendapatan = KasMasuk.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    total_beban = KasKeluar.objects.aggregate(total=Sum('jumlah'))['total'] or 0
    laba_rugi = total_pendapatan - total_beban

    wb = Workbook()
    ws = wb.active
    ws.title = "Laba Rugi"

    ws.append(['Keterangan', 'Jumlah'])
    ws.append(['Pendapatan / Kas Masuk', float(total_pendapatan)])
    ws.append(['Beban / Kas Keluar', float(total_beban)])
    ws.append(['Laba / Rugi', float(laba_rugi)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=laba_rugi.xlsx'

    wb.save(response)
    return response